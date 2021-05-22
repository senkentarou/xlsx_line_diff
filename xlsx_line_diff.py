# coding: utf-8

import sys
import openpyxl

from argparse import ArgumentParser
from difflib import ndiff

def args():
    argparser = ArgumentParser(description='xlsx line diff')
    argparser.add_argument('xlsx_from_path', type=str)
    argparser.add_argument('xlsx_to_path', type=str)
    argparser.add_argument('--page-name', type=str)

    return argparser.parse_args()

def csv_lines(page):
    return [','.join(str(r.value) for r in row) for row in page.rows]

def print_csv_contents_diff(line_a, line_b):
    for a, b in zip(line_a.split(','), line_b.split(',')):
        if a != b:
            print('{} <=> {}'.format(cyan_text(a), red_text(b)))

def cyan_text(message):
    return '\033[36m{}\033[0m'.format(message)

def red_text(message):
    return '\033[31m{}\033[0m'.format(message)

def main(args):
    xlsx_from_path, xlsx_to_path = args.xlsx_from_path, args.xlsx_to_path

    print('- Loading {} ...'.format(xlsx_from_path))
    xf_wb = openpyxl.load_workbook(xlsx_from_path)
    print('+ Loading {} ...'.format(xlsx_to_path))
    xt_wb = openpyxl.load_workbook(xlsx_to_path)

    # print('Selecting page ...')
    page_name = args.page_name
    if page_name is None:
        xf_page, xt_page = xf_wb.active, xt_wb.active
    elif page_name in xf_wb.sheetnames and page_name in xt_wb.sheetnames:
        xf_page, xt_page = xf_wb[page_name], xt_wb[page_name]
    else:
        print('--page-name \'{}\' does not exist in each file'.format(page_name))
        return 1

    # print('Creating csv lines ...')
    xf_lines, xt_lines = csv_lines(xf_page), csv_lines(xt_page)

    print('---------- Diff lines! ----------')
    # below is yabai algorithm..
    former_diff_line = ' '
    is_marked_hunk = False
    diff_line_idx = 1
    for line in ndiff(xf_lines, xt_lines):
        if line[0] not in ['+', '-', '?']:
            continue
        if line[0] == '?':
            if former_diff_line[0] == '-':
                is_marked_hunk = True
            continue

        print('{}: {}'.format(diff_line_idx, line))

        if is_marked_hunk:
            print_csv_contents_diff(former_diff_line[1:], line[1:])
            is_marked_hunk = False

        former_diff_line = line
        diff_line_idx += 1

    print('----------    Done.    ----------')
    return 0

if __name__ == '__main__':
    sys.exit(main(args()))
